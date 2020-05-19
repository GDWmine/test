import openpyxl


# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('template.xlsx')
sheets = wb.sheetnames
print(sheets, type(sheets))
# 创建一个表单
mySheet = wb.create_sheet('mySheet111')
print(wb.sheetnames)
 
# 获取指定的表单
sheet3 = wb.get_sheet_by_name('Sheet3')
sheet4 = wb['mySheet111']
for sheet in wb:
    print(sheet.title)

ws = wb.active 
print(ws)
print(ws['A1']) # 获取A列的第一个对象
print(ws['B1'])
print(ws['A1'].value)
print(ws['B1'].value)